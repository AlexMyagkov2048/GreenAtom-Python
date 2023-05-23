import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
import re
import openpyxl
from datetime import date
from openpyxl.styles import Alignment
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from dotenv import load_dotenv
import os

today = date.today()
formatted_date = today.strftime("%d-%m-%Y")

#Запускаем Браузер
options = Options()
options.add_experimental_option("excludeSwitches", ['enable-logging'])
options.add_argument("user-agent=[Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36]")
service = Service('C:\\Users\\User\\PycharmProjects\\resource\\chromedriver.exe')
driver = webdriver.Chrome(service=service, options=options)
print("Начинаем работу парсера")

url = 'https://www.moex.com/'
search = '//input[@id="moex-search-input"]'


usd_rub = '/html/body/div[3]/div[3]/div/div/div[1]/div[2]/div/div/div/div[2]/form/div[2]/div[3]/select'
count_rows = '/html/body/span[9]'
table = '//table[@style="border: solid 1px #666666; border-collapse:collapse;"]'

#Открываем браузер на весь экран
driver.get(url)
get_url = driver.current_url
print("Current url:", get_url)
driver.maximize_window()

#Переходим по ссылке Индикативные курсы
search_field = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, search)))
search_field.click()
search_field.send_keys("Индикативные курсы валют")
time.sleep(5)
elements = WebDriverWait(driver, 20).until(EC.visibility_of_all_elements_located((By.XPATH, '//a[@itemprop="item"]')))
Indicative_сourses = elements[2]
third_element_text = Indicative_сourses.text
try:
    assert third_element_text == "Индикативные курсы"
    print(f"Клик по ссылке: {third_element_text}")
    Indicative_сourses.click()
    accept_terms = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, '//a[@data-dismiss="modal"]')))
    accept_terms.click()
    time.sleep(5)
except AssertionError:
    print("Текст элемента не соответствует")

# Переходим на xml страницу
xml = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//div[contains(@class, 'col-md-48') and contains(@class, 'text-center') and contains(@class, 'margin-tb-10')]")))
xml.click()
get_url = driver.current_url
attempts = 0
while attempts < 3:
    try:
        get_url = driver.current_url
        match = re.search(r'currency=(\w+/\w+)', get_url)
        currency = match.group(1)
        print(f"Проверка  валютной пары: {currency}")
        assert currency == "USD/RUB"
        print("Валютная пара USD/RUB")
        break
    except (AttributeError, AssertionError, Exception):
        print('Неверная валютная пара  или URL не найден')
        get_url = re.sub(r'currency=([^/]*)', r'currency=RUB', get_url)
        driver.get(get_url)
        attempts += 1
        if attempts == 3:
            print("Не удалось получить  валютную пару. Осуществите проверку кода.")
            exit()
else:
    print("Не удалось получить  валютную пару. Осуществите проверку кода.")


get_url = re.sub(r'currency=([^/]*)', r'currency=JPY', get_url)
count_rows_value = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, count_rows)))
count_rows_value_text = count_rows_value.text
print(count_rows_value_text)

table_1 = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, table)))
table_1_value = table_1.text


excel_file = f"result_{formatted_date}.xlsx"

# Открываем файл Excel
workbook = openpyxl.Workbook()
# Получаем активный лист
worksheet = workbook.active
worksheet['A1'] = 'Дата USD/RUB'
worksheet['B1'] = 'Курс USD/RUB'
worksheet['C1'] = 'Время USD/RUB'
worksheet['G1'] = 'Результат'


#Разделяем строки
table_rows = table_1_value.replace('\r','').split('\n')
for row in range(2, len(table_rows), 2):
#Задаём ячейкам финансовый формат
    table_cols = table_rows[row].split()
    date = table_cols[1]
    time = table_cols[2]
    value = table_cols[3]
    worksheet.cell(row=row, column=2, value=value).number_format = '# ##0.00' + ' ' + u'\u20BD'
    worksheet.cell(row=row, column=5, value=value).number_format = '# ##0.00' + ' ' + u'\u20BD'
    worksheet.cell(row=row, column=7).number_format = '# ##0.00' + ' ' + u'\u20BD'

#Заполняем колонки
    value_num = float(value)  # преобразуем строку в число
    worksheet.cell(row=row, column=1, value=date)
    worksheet.cell(row=row, column=2, value=value_num)
    worksheet.cell(row=row, column=3, value=time)


# Сохраняем файл
workbook.save(excel_file)

print(f"url для JPY/RUB: {get_url}")
driver.get(get_url)
table_2 = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, table)))
table_2_value = table_2.text
driver.close()

# Открываем файл Excel
workbook = openpyxl.load_workbook(excel_file)
# Получаем активный лист
worksheet = workbook.active
worksheet['D1'] = 'Дата JPY/RUB'
worksheet['E1'] = 'Курс JPY/RUB'
worksheet['F1'] = 'Время JPY/RUB'

# Разделяем строки и записсываем в диапазон ячеек
table_rows = table_2_value.replace('\r','').split('\n')
for row in range(2, len(table_rows), 2):
    table_cols = table_rows[row].split()
    date = table_cols[1]
    time = table_cols[2]
    value = table_cols[3]
    value_num = float(value)
    worksheet.cell(row=row, column=4, value=date)
    worksheet.cell(row=row, column=5, value=value_num)
    worksheet.cell(row=row, column=6, value=time)



# Выравнивание содержимого ячеек по центру
for row in worksheet.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Автоширина
for col in worksheet.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    worksheet.column_dimensions[col_letter].width = adjusted_width


worksheet = workbook.active

# Удаляем пустые строки
max_row = worksheet.max_row
for row in range(max_row, 1, -1):
    if not any(cell.value for cell in worksheet[row]):
        worksheet.delete_rows(row)

num_rows = worksheet.max_row
for row in range(2, num_rows + 1):
    worksheet.cell(row=row, column=7).value = f"=B{row}/E{row}"
print(num_rows)

# Сохраняем файл
workbook.save(excel_file)

forms = ['строка', 'строки', 'строк']
def declension(num, forms):
    if num % 100 in [11, 12, 13, 14]:
        return f'{num} {forms[2]}'
    elif num % 10 == 1:
        return f'{num} {forms[0]}'
    elif num % 10 in [2, 3, 4]:
        return f'{num} {forms[1]}'
    else:
        return f'{num} {forms[2]}'


def send_mail():
    # Указываем параметры для подключения к серверу SMTP
    load_dotenv()
    smtp_host = 'smtp.yandex.ru'
    smtp_port = 587
    smtp_user = 'myagkoff.m89@yandex.ru'
    smtp_password = os.getenv('PASSWORD')

    # Создаем объект MIMEMultipart для добавления текста и вложения
    msg = MIMEMultipart()
    msg['From'] = 'myagkoff.m89@yandex.ru'
    msg['To'] = 'myagkoff.m89@yandex.ru'
    msg['Subject'] = 'Тестовое письмо'

    # Добавляем текст сообщения
    body = text
    msg.attach(MIMEText(body, 'plain'))

    # Добавляем вложение
    with open(f'{excel_file}', 'rb') as f:
        attach = MIMEApplication(f.read(), _subtype='xlsx')
        attach.add_header('Content-Disposition', 'attachment', filename=f'{excel_file}')
        msg.attach(attach)

    # Отправляем письмо
    with smtplib.SMTP(smtp_host, smtp_port) as server:
        server.starttls()
        server.login(smtp_user, smtp_password)
        server.send_message(msg)

    print(" Письмо отправлено")

text = f"В таблице {declension(num_rows, forms)}"
declension (num_rows, forms)
print(text)
send_mail()